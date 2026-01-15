import os
import time

import openpyxl

# get the date in yyyymmmdd format

date = time.strftime("%Y%m%d", time.localtime())

out_dir = f"data/proc/tubulin_alignments/{date}_tubulin_protein_seqs_{date}"

# check if the directory exists, if not create it
if not os.path.exists(out_dir):
    print(f"Creating output directory: {out_dir}")
    os.makedirs(out_dir)


ce_prot_file = "data/blast_data/20230324_blast/c_elegans/ce_all_prot_flat.fa"
cb_prot_file = "data/blast_data/20230324_blast/c_briggsae/cb_all_prot_flat.fa"
ct_prot_file = "data/blast_data/20230324_blast/c_tropicalis/ct_all_prot_flat.fa"
pp_prot_file = "data/pacificus_bts/raw/P_pacifcus_Beta_tubulin_genes_amino_acid_sequences.xlsx"


# out_file = 'test_pull.fa'
# out_file_cb = 'test_pull_cb.fa'
# out_file_ce = 'test_pull_ce.fa'


# function that will pull the line with the sequence id and the next line with the sequence
def pull_protein_sequence(protein_id, prot_file, out_file):
    print(f"Pulling protein sequence for ID: {protein_id} from file: {prot_file}")
    with open(prot_file, "r") as f:
        lines = f.readlines()
        print(f"Total lines read from {prot_file}: {len(lines)}")
        for i, line in enumerate(lines):
            if protein_id in line:
                print(f"Found protein ID {protein_id} at line {i}")
                with open(out_file, "a") as out_f:
                    out_f.write(line)
                    out_f.write(lines[i + 1])
                print(f"Protein sequence written to {out_file}")
                break
        else:
            print(f"Protein ID {protein_id} not found in {prot_file}")


# pull_protein_sequence('QX1410.13336.1', cb_prot_file, out_file_cb)
# pull_protein_sequence('C54C6.2.1', ce_prot_file, out_file_ce)

#### ben_1 ####
# ben_1 = {
#     'cb': 'QX1410.13336.1',
#     'ce': 'C54C6.2.1',
#     'ct': 'NIC58.15504.1'
# }


def pp_pull_and_rename_protein_sequence(protein_id, formal_name, prot_file, out_file):
    """Pull protein sequence from P. pacificus Excel file and write to output file.

    Args:
        protein_id: Gene ID to search for in the 'Gene ID' column
        formal_name: Name to use in the output fasta header
        prot_file: Path to the Excel file containing P. pacificus sequences
        out_file: Path to output fasta file
    """
    print(f"Pulling P.p protein sequence for ID: {protein_id} from file: {prot_file}")
    wb = openpyxl.load_workbook(prot_file)
    ws = wb.active

    # Find the column indices (assuming header row is row 1)
    header = [cell.value for cell in ws[1]]
    gene_id_col = header.index("Gene ID") + 1  # openpyxl uses 1-based indexing
    aa_seq_col = header.index("Amino Acid sequence") + 1

    # Search for the protein_id in the Gene ID column
    for row in ws.iter_rows(min_row=2):  # Skip header row
        gene_id = row[gene_id_col - 1].value  # Convert to 0-based for row access
        if gene_id == protein_id:
            aa_sequence = row[aa_seq_col - 1].value
            print(f"Found protein ID {protein_id}")
            with open(out_file, "a") as out_f:
                out_f.write(f">{formal_name}\n")
                out_f.write(f"{aa_sequence}\n")
            print(f"Protein sequence written to {out_file}")
            wb.close()
            return

    print(f"Protein ID {protein_id} not found in {prot_file}")
    wb.close()


def pull_and_rename_protein_sequence(protein_id, formal_name, prot_file, out_file):
    print(f"Pulling protein sequence for ID: {protein_id} from file: {prot_file}")
    with open(prot_file, "r") as f:
        lines = f.readlines()
        print(f"Total lines read from {prot_file}: {len(lines)}")
        for i, line in enumerate(lines):
            if protein_id in line:
                print(f"Found protein ID {protein_id} at line {i}")
                with open(out_file, "a") as out_f:
                    out_f.write(
                        f">{formal_name}\n"
                    )  # Write the formal name as the new header
                    out_f.write(lines[i + 1])  # Write the sequence


def process_protein_sequences(
    protein_dict, ce_prot_file, cb_prot_file, ct_prot_file, pp_prot_file, out_file
):
    for key, protein_info in protein_dict.items():
        protein_id = protein_info[0]  # Extract the transcript ID
        formal_name = protein_info[1]  # Extract the formal name
        if key == "ce":
            pull_and_rename_protein_sequence(
                protein_id, formal_name, ce_prot_file, out_file
            )
        elif key == "cb":
            pull_and_rename_protein_sequence(
                protein_id, formal_name, cb_prot_file, out_file
            )
        elif key == "ct":
            pull_and_rename_protein_sequence(
                protein_id, formal_name, ct_prot_file, out_file
            )
        elif key == "pp":
            pp_pull_and_rename_protein_sequence(
                protein_id, formal_name, pp_prot_file, out_file
            )
        else:
            print(f"Invalid species key {key}")


# use a list for each entry to include transcript ID and the formal name
ben_1 = {
    "ce": ["C54C6.2.1", "CE_ben-1"],
    "cb": ["QX1410.13336.1", "CB_ben-1"],
    "ct": ["NIC58.15504.1", "CT_ben-1"],
}

out_file = f"{out_dir}/ben_1.fa"

# check if the file exists and contains data, if so, delete it
if os.path.exists(out_file):
    print(f"Deleting existing file: {out_file}")
    os.remove(out_file)

process_protein_sequences(ben_1, ce_prot_file, cb_prot_file, ct_prot_file, out_file)

tbb_1 = {
    "ce": ["K01G5.7.1", "CE_tbb-1"],
    "cb": ["QX1410.12556.1", "CB_tbb-1"],
    "ct": ["NIC58.14724.4", "CT_tbb-1"],
}

out_file = f"{out_dir}/tbb_1.fa"

# check if the file exists and contains data, if so, delete it
if os.path.exists(out_file):
    print(f"Deleting existing file: {out_file}")
    os.remove(out_file)

process_protein_sequences(tbb_1, ce_prot_file, cb_prot_file, ct_prot_file, out_file)


tbb_2 = {
    "ce": ["C36E8.5.1", "CE_tbb-2"],
    "cb": ["QX1410.14085.1", "CB_tbb-2"],
    "ct": ["NIC58.15977.2", "CT_tbb-2"],
}

out_file = f"{out_dir}/tbb_2.fa"

# check if the file exists and contains data, if so, delete it
if os.path.exists(out_file):
    print(f"Deleting existing file: {out_file}")
    os.remove(out_file)

process_protein_sequences(tbb_2, ce_prot_file, cb_prot_file, ct_prot_file, out_file)

mec_7 = {
    "ce": ["ZK154.3.1", "CE_mec-7"],
    "cb": ["QX1410.17185.1", "CB_mec-7"],
    "ct": ["NIC58.18736.4", "CT_mec-7"],
}

out_file = f"{out_dir}/mec_7.fa"

# check if the file exists and contains data, if so, delete it
if os.path.exists(out_file):
    print(f"Deleting existing file: {out_file}")
    os.remove(out_file)

process_protein_sequences(mec_7, ce_prot_file, cb_prot_file, ct_prot_file, out_file)

tbb_4 = {
    "ce": ["B0272.1.1", "CE_tbb-4"],
    "cb": ["QX1410.15075.1", "CB_tbb-4"],
    "ct": ["NIC58.17262.1", "CT_tbb-4"],
}

out_file = f"{out_dir}/tbb_4.fa"

# check if the file exists and contains data, if so, delete it
if os.path.exists(out_file):
    print(f"Deleting existing file: {out_file}")
    os.remove(out_file)

process_protein_sequences(tbb_4, ce_prot_file, cb_prot_file, ct_prot_file, out_file)
